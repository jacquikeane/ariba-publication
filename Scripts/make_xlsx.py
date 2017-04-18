#!/usr/bin/env python3
from openpyxl.utils import get_column_letter
import openpyxl


filenames_to_load = [
    'E_faecium/summary.supplementary_table.tsv',
    'Time_and_memory/resources.tsv',
    'S_sonnei/shigella_supplementary_holt2012.tsv',
    'S_sonnei/Ref/ref_data.in.tsv',
    'S_sonnei/summary.calls.tsv',
    'N_gonorrhoeae/fig5.mic_data.csv',
    'N_gonorrhoeae/suppl_table_gono_antibios.xlsx',
    'S_sonnei/summary.amr_genes.tsv',
]


# can't copy sheet between workbooks, so read the xlsx file, which
# has the one sheet that isn't made from a TSV. Then add in
# a new sheet per TSV file.
wb = openpyxl.load_workbook('N_gonorrhoeae/suppl_table_gono_antibios.xlsx')
print(wb.sheetnames)
found_xlxs = False

for sheet_index in range(len(filenames_to_load)):
    worksheet_name = 'Supplementary Table ' + str(sheet_index + 1)
    filename = filenames_to_load[sheet_index]

    if filename.endswith('.xlsx'):
        found_xlxs = True
        for sheet in wb:
            if not sheet.title.startswith('Suppl'):
                sheet.title = worksheet_name
        continue

    if found_xlxs:
        worksheet = wb.create_sheet(worksheet_name)
    else:
        worksheet = wb.create_sheet(worksheet_name, len(wb.sheetnames) - 1)

    delim = ',' if filename.endswith('csv') else '\t'
    col_widths = []

    with open(filename) as f:
        for line in f:
            data = line.rstrip().split(delim)
            for i in range(len(data)):
                try:
                    data[i] = float(data[i])
                except:
                    if data[i].endswith('%'):
                        try:
                            data[i] = float(data[i].rstrip('%'))
                        except:
                            pass

            if len(col_widths) == 0:
                col_widths = [len(str(data[i])) for i in range(len(data))]

            for i in data:
                col_widths = [max(col_widths[i], len(str(data[i]))) for i in range(len(data))]
            worksheet.append(data)

    for i in range(len(col_widths)):
        worksheet.column_dimensions[get_column_letter(i+1)].width = col_widths[i]

    worksheet.freeze_panes = 'A2'

wb.save('supplementary_tables.xlsx')
